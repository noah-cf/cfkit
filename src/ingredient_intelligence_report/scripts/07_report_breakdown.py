import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os

script_dir = os.path.dirname(os.path.abspath(__file__))
input_file = os.path.join(script_dir, '../../../data/ingredient_intelligence_reports/07_report_breakdown_input.xlsx')
output_file = os.path.join(script_dir, '../../../data/ingredient_intelligence_reports/07_report_breakdown_output.xlsx')

df = pd.read_excel(input_file)

if 'count' not in df.columns:
    df['count'] = 1
else:
    df['count'] = df['count'].fillna(1)

df['Chemical Count'] = df.groupby('Final CAS')['count'].transform('sum')

wb = load_workbook(filename=input_file)
ws_source = wb.active

ws_source.title = "Source+Data"

conditions = {
    "Low Rated": ["F", "D"],
    "Data Gaps": "?",
    "Unique Low Rated": ["F", "D"],
    "Unique Data Gaps": "?"
}

for sheet_name, condition in conditions.items():
    ws = wb.create_sheet(title=sheet_name)

    if isinstance(condition, list):
        df_filtered = df[df['Hazard Band'].isin(condition)]
    else:
        df_filtered = df[df['Hazard Band'] == condition]

    if sheet_name.startswith("Unique"):
        df_filtered = df_filtered.drop_duplicates(subset='Final CAS')
        cols_to_drop = df_filtered.columns.tolist()[:df_filtered.columns.tolist().index('Chemical Count')]
        df_filtered = df_filtered.drop(columns=cols_to_drop)

    df_filtered = df_filtered.sort_values(by='Chemical Count', ascending=False)

    rows = dataframe_to_rows(df_filtered, index=False, header=True)

    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

for sheet_name in ["Unidentified", "Unique Unidentified"]:
    ws = wb.create_sheet(title=sheet_name)

    df_filtered = df[df['Ingredient Name'].isna()]

    if sheet_name.startswith("Unique"):
        df_filtered = df_filtered.drop_duplicates(subset='Final CAS')
        cols_to_drop = df_filtered.columns.tolist()[:df_filtered.columns.tolist().index('Chemical Count')]
        df_filtered = df_filtered.drop(columns=cols_to_drop)

    df_filtered = df_filtered.sort_values(by='Chemical Count', ascending=False)

    rows = dataframe_to_rows(df_filtered, index=False, header=True)

    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

ws = wb.create_sheet(title="Unique Source+Data")

df_filtered = df.drop_duplicates(subset='Final CAS')
cols_to_drop = df_filtered.columns.tolist()[:df_filtered.columns.tolist().index('Chemical Count')]
df_filtered = df_filtered.drop(columns=cols_to_drop)

df_filtered = df_filtered.sort_values(by='Chemical Count', ascending=False)

rows = dataframe_to_rows(df_filtered, index=False, header=True)

for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

sheet_order = ["Source+Data",
               "Low Rated",
               "Data Gaps",
               "Unidentified",
               "Unique Source+Data",
               "Unique Low Rated",
               "Unique Data Gaps",
               "Unique Unidentified"
]

wb._sheets = [wb[sheet_name] for sheet_name in sheet_order]

wb.save(output_file)
