import re
import numpy as np
import pandas as pd
import os
from openpyxl import load_workbook

script_dir = os.path.dirname(os.path.abspath(__file__))
input_file = os.path.join(script_dir, '../../../data/ingredient_intelligence_reports/01_cas_cleaning_input.xlsx')
output_file = os.path.join(script_dir, '../../../data/ingredient_intelligence_reports/01_cas_cleaning_output.xlsx')
cas_column = 'cas'

def verify_cas(cas):
    if not isinstance(cas, str):
        cas = str(cas)

    if cas.startswith('0'):
        return False

    cas = cas.replace('-', '')
    if not cas.isnumeric():
        return False
    reverse_cas = cas[::-1]
    checksum = int(reverse_cas[0])
    rest = reverse_cas[1:]
    expected_checksum = sum(int(digit) * i for i, digit in enumerate(rest, start=1)) % 10
    return checksum == expected_checksum

def correct_cas(cas):
    cas = cas.lstrip('0')
    cas = re.sub(r'\s*-\s*', '-', cas)

    if ' 00:00:00' in cas:
        cas_parts = cas.split(' ')[0].split('-')
        cas_parts[-1] = cas_parts[-1].lstrip('0')
        cas = '-'.join(cas_parts)
    else:
        cas_parts = cas.split('-')
        if len(cas_parts) == 3:
            cas_parts[0] = cas_parts[0].lstrip('0')
            cas_parts[-1] = cas_parts[-1].lstrip('0')
            cas = '-'.join(cas_parts)
    return cas

def final_cas(row, cas_column):
    if row['Valid Original']:
        return row[cas_column]
    elif row['Valid Corrected']:
        return row['Corrected CAS']
    elif str(row['Corrected CAS']).lower() == 'nan':
        return 'no_cas'
    else:
        return 'invalid_cas'

def analyze_cas(file_path, cas_column):
    dtype = {cas_column: str}

    if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
        data = pd.read_excel(file_path, dtype=dtype)
    elif file_path.endswith('.csv'):
        data = pd.read_csv(file_path, dtype=dtype)
    else:
        raise ValueError("Invalid file format. Please provide an Excel (.xlsx, .xls) or CSV (.csv) file.")

    cas_df = data[cas_column].astype(str).to_frame()
    cas_df['Valid Original'] = cas_df[cas_column].apply(verify_cas)
    cas_df['Corrected CAS'] = cas_df.loc[~cas_df['Valid Original'], cas_column].apply(correct_cas)
    cas_df['Valid Corrected'] = cas_df['Corrected CAS'].apply(verify_cas)
    cas_df['Final Valid'] = cas_df['Valid Original'] | cas_df['Valid Corrected']
    cas_df['Final CAS'] = cas_df.apply(final_cas, args=(cas_column,), axis=1)

    data = pd.concat([data, cas_df[[
        'Valid Original',
        'Corrected CAS',
        'Valid Corrected',
        'Final Valid',
        'Final CAS'
    ]]], axis=1)

    return data

result = analyze_cas(input_file, cas_column)

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    result.to_excel(writer, index=False, sheet_name='Sheet1')

wb = load_workbook(output_file)
ws = wb.active
for row in ws.iter_rows(min_row=2, max_row=len(result) + 1):
    cas_cell = row[0]
    cas_cell.number_format = '@'

wb.save(output_file)
