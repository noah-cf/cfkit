from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
import os

script_dir = os.path.dirname(os.path.abspath(__file__))
input_file = os.path.join(script_dir, '../../../data/ingredient_intelligence_reports/08_report_formatting_input.xlsx')
output_file = os.path.join(script_dir, '../../../data/ingredient_intelligence_reports/08_report_formatting_output.xlsx')

wb = load_workbook(filename=input_file)

fills = {
    'A': PatternFill(start_color="006400", end_color="006400", fill_type="solid"),
    'B': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
    'B_C': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
    'C': PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"),
    'D': PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
    'F': PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),
    '?': PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),
    'IP': PatternFill(start_color="DA70D6", end_color="DA70D6", fill_type="solid"),
    'Assessment in Progress': PatternFill(start_color="DA70D6", end_color="DA70D6", fill_type="solid"),
    'Full Chemical Hazard Assessment': PatternFill(start_color="006400", end_color="006400", fill_type="solid"),
    'Full Green Circle': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
    'Half Green Circle': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
    'Yellow Triangle': PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"),
    'Tentative': PatternFill(start_color="DA70D6", end_color="DA70D6", fill_type="solid"),
    'BM-3': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
    'BM-2': PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"),
}

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for ws in wb.worksheets:
    for col_idx, cell in enumerate(ws[1], start=1):
        col_letter = get_column_letter(col_idx)
        column_name = cell.value

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.fill = fills.get(cell.value, PatternFill())

        if column_name == 'Final CAS':
            ws.column_dimensions[col_letter].width = 12
            for cell in ws[col_letter]:
                cell.alignment = Alignment(horizontal='center')
                cell.number_format = '@'
        elif column_name == 'Additional CAS':
            ws.column_dimensions[col_letter].width = 12
            for cell in ws[col_letter]:
                cell.alignment = Alignment(horizontal='left')
        elif column_name == 'EC Number':
            ws.column_dimensions[col_letter].width = 12
            for cell in ws[col_letter]:
                cell.alignment = Alignment(horizontal='left')
        elif column_name == 'Ingredient Name':
            ws.column_dimensions[col_letter].width = 40
            for cell in ws[col_letter]:
                cell.alignment = Alignment(horizontal='left')
        elif column_name == 'INCI':
            ws.column_dimensions[col_letter].width = 30
            for cell in ws[col_letter]:
                cell.alignment = Alignment(horizontal='left')
        elif column_name == 'Hazard Band':
            ws.column_dimensions[col_letter].width = 12
            for cell in ws[col_letter]:
                cell.alignment = Alignment(horizontal='center')
        elif column_name == 'C2C Score':
            ws.column_dimensions[col_letter].width = 12
            for cell in ws[col_letter]:
                cell.alignment = Alignment(horizontal='center')
        elif column_name == 'ChemFORWARD Status':
            ws.column_dimensions[col_letter].width = 30
            for cell in ws[col_letter]:
                cell.alignment = Alignment(horizontal='center')
        elif column_name == 'SCIL Status':
            ws.column_dimensions[col_letter].width = 14
            for cell in ws[col_letter]:
                cell.alignment = Alignment(horizontal='center')
        elif column_name == 'TCO Status':
            ws.column_dimensions[col_letter].width = 14
            for cell in ws[col_letter]:
                cell.alignment = Alignment(horizontal='center')
        elif column_name in ['replaced?',
                             'original_Final CAS',
                             'cas',
                             'Count',
                             'count',
                             'Valid Original',
                             'Corrected CAS',
                             'Final Valid',
                             'Chemical Count'
            ]:
            for cell in ws[col_letter]:
                cell.alignment = Alignment(horizontal='center')

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.border = thin_border

wb.save(output_file)
