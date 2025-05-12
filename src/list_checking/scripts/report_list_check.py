import os
import pandas as pd
from datetime import datetime
import shutil
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.cell import Cell

def get_input_file():
    """Get the input file from the data directory."""
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../../data/list_checking/report_list_check"))
    files = [f for f in os.listdir(base_dir) if f.endswith(('.xlsx', '.csv')) and os.path.isfile(os.path.join(base_dir, f))]
    
    if not files:
        raise FileNotFoundError(f"No xlsx or csv files found in {base_dir}")

    if len(files) == 1:
        return os.path.join(base_dir, files[0])
    
    print("\nMultiple files found. Please select one:")
    for i, file in enumerate(files, 1):
        print(f"{i}. {file}")
    
    while True:
        try:
            choice = int(input("Enter the number of the file to process: "))
            if 1 <= choice <= len(files):
                return os.path.join(base_dir, files[choice - 1])
            else:
                print("Invalid choice. Please try again.")
        except ValueError:
            print("Please enter a valid number.")

def show_data_preview(df):
    """Show a preview of the dataframe to help user identify CAS column."""
    print("\nData preview (first 5 rows):")
    print(df.head().to_string())
    print("\nColumn names:")
    for i, col in enumerate(df.columns, 1):
        print(f"{i}. {col}")

def get_cas_column(df):
    """Get the CAS column from user input."""
    show_data_preview(df)
    
    while True:
        col_input = input("\nEnter the column name or number containing CAS numbers: ")
        
        try:
            col_num = int(col_input)
            if 1 <= col_num <= len(df.columns):
                return df.columns[col_num - 1]
            else:
                print("Invalid column number. Please try again.")
                continue
        except ValueError:
            # Not a number, try as column name
            if col_input in df.columns:
                return col_input
            else:
                print(f"Column '{col_input}' not found. Please try again.")

def load_list_file(file_path):
    """Load a list file and extract the list name, sublist name, and CAS numbers."""
    df = pd.read_excel(file_path, header=None)
    
    list_name = None
    sublist_name = None
    data_start_row = None
    
    for idx, row in df.iterrows():
        if row[0] == 'list:':
            list_name = row[1]
        elif row[0] == 'sublist:':
            sublist_name = row[1]
        elif 'CASRN' in row.values.tolist():
            data_start_row = idx
            break
    
    if list_name is None or data_start_row is None:
        raise ValueError(f"Could not parse list file: {file_path}")
    
    if sublist_name:
        combined_name = f"{list_name} [{sublist_name}]"
    else:
        combined_name = list_name
    
    if len(combined_name) > 100:
        if sublist_name:
            shortened_list = list_name[:50] + "..."
            combined_name = f"{shortened_list} [{sublist_name}]"
            if len(combined_name) > 100:
                shortened_sublist = sublist_name[:30] + "..."
                combined_name = f"{shortened_list} [{shortened_sublist}]"
        else:
            combined_name = list_name[:97] + "..."
    
    df_data = pd.read_excel(file_path, skiprows=data_start_row)
    cas_numbers = set(df_data['CASRN'].dropna().astype(str))
    
    return combined_name, cas_numbers

def load_all_lists(base_dir):
    """Load all list files from regulatory and non_regulatory directories."""
    lists_data = {'regulatory': {}, 'non_regulatory': {}}
    
    for list_type in ['regulatory', 'non_regulatory']:
        dir_path = os.path.join(base_dir, list_type)
        if not os.path.exists(dir_path):
            continue
            
        for file in os.listdir(dir_path):
            if file.endswith('.xlsx'):
                file_path = os.path.join(dir_path, file)
                try:
                    list_name, cas_numbers = load_list_file(file_path)
                    lists_data[list_type][list_name] = cas_numbers
                except Exception as e:
                    print(f"Error loading {file}: {e}")
                    continue
    
    return lists_data

def create_output_directory(base_dir):
    """Create timestamped output directory structure."""
    timestamp = datetime.now().strftime('%Y_%m_%d_%H_%M_%S')
    output_dir = os.path.join(os.path.dirname(base_dir), f"{timestamp}_report_list_check_output")
    
    os.makedirs(os.path.join(output_dir, 'input'), exist_ok=True)
    os.makedirs(os.path.join(output_dir, 'regulatory'), exist_ok=True)
    os.makedirs(os.path.join(output_dir, 'non_regulatory'), exist_ok=True)
    
    return output_dir

def copy_files_to_output(input_file, lists_dir, output_dir):
    """Copy input and list files to output directory."""
    shutil.copy2(input_file, os.path.join(output_dir, 'input'))
    
    for list_type in ['regulatory', 'non_regulatory']:
        src_dir = os.path.join(lists_dir, list_type)
        dst_dir = os.path.join(output_dir, list_type)
        
        if os.path.exists(src_dir):
            for file in os.listdir(src_dir):
                if file.endswith('.xlsx'):
                    shutil.copy2(os.path.join(src_dir, file), dst_dir)

def process_cas_matching(df, cas_column, lists_data):
    """Process CAS matching and add columns for each list."""
    unique_cas = df[cas_column].dropna().astype(str).unique()
    
    all_lists = []
    for list_type in ['regulatory', 'non_regulatory']:
        for list_name in sorted(lists_data[list_type].keys()):
            all_lists.append((list_type, list_name))
            df[list_name] = ''
    
    print("\nProcessing CAS numbers...")
    for idx, row in tqdm(df.iterrows(), total=len(df), desc="Matching CAS numbers"):
        cas = str(row[cas_column]).strip()
        if pd.isna(cas) or cas == 'nan':
            continue
            
        for list_type, list_name in all_lists:
            if cas in lists_data[list_type][list_name]:
                df.at[idx, list_name] = cas
    
    return df

def save_output_excel(df, output_file, cas_column, list_columns):
    """Save output Excel file with proper text formatting and center alignment for CAS columns."""
    df.to_excel(output_file, index=False)
    
    wb = load_workbook(output_file)
    ws = wb.active
    
    columns_to_format = [cas_column] + list_columns
    col_indices = []
    for col in columns_to_format:
        if col in df.columns:
            col_indices.append(df.columns.get_loc(col) + 1)
    
    from openpyxl.styles import Alignment
    
    for col_idx in col_indices:
        header_cell = ws.cell(row=1, column=col_idx)
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row in range(2, len(df) + 2):
            cell = ws.cell(row=row, column=col_idx)
            cell.number_format = '@'
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    wb.save(output_file)

def main():
    try:
        input_file = get_input_file()
        input_filename = os.path.basename(input_file)
        
        if input_file.endswith('.xlsx'):
            df = pd.read_excel(input_file)
        else:
            df = pd.read_csv(input_file)
        
        cas_column = get_cas_column(df)
        lists_dir = os.path.join(os.path.dirname(input_file))
        lists_data = load_all_lists(lists_dir)
        output_dir = create_output_directory(lists_dir)
        
        print("\nCopying files to output directory...")
        copy_files_to_output(input_file, lists_dir, output_dir)
        
        df = process_cas_matching(df, cas_column, lists_data)
        
        list_columns = []
        for list_type in ['regulatory', 'non_regulatory']:
            list_columns.extend(sorted(lists_data[list_type].keys()))
        
        output_filename = f"{os.path.splitext(input_filename)[0]}_checked.xlsx"
        output_file = os.path.join(output_dir, output_filename)
        
        print("\nSaving output file...")
        save_output_excel(df, output_file, cas_column, list_columns)
        
        print(f"\nProcessing complete! Output saved to:\n{output_file}")
        
    except Exception as e:
        print(f"\nError: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()